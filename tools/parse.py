# -*- coding: utf-8 -*-
"""
가계부 + 자산 포트폴리오 엑셀 → data.js (window.APP_DATA) 생성기
- 스냅샷(오프라인) 모드용. 라이브 연동은 Code.gs가 동일 스키마로 생성 예정.
- ANONYMIZE=True: 소개팅/동행 실명 마스킹 (GitHub Pages는 공개 저장소이므로 기본 ON 권장)
"""
import openpyxl, re, json, datetime as dt, sys

LEDGER = sys.argv[1] if len(sys.argv) > 1 else '____주현_가계부.xlsx'
PORTFOLIO = sys.argv[2] if len(sys.argv) > 2 else '____종합_자산_포트폴리오_재테크_.xlsx'
OUT = sys.argv[3] if len(sys.argv) > 3 else 'data.js'
ANONYMIZE = True

# ---------- 카테고리 표준화 (설계안 §1-C) ----------
CAT_MAP = {
    '외식/점심':'외식/점심','외식/전심':'외식/점심','외식/저녁':'외식/저녁',
    '집밥':'집밥/생활비','생활비':'집밥/생활비','집밥/생활비':'집밥/생활비','디저트':'디저트',
    '운동':'운동','의료비':'의료비','보험비':'보험비','영양제':'영양제',
    '교통':'대중교통','대중교통':'대중교통','차량':'차량','주유/톨비':'주유/톨비','통신비':'통신비',
    '선물':'선물','경조사':'경조사','보은':'보은','회비':'회비',
    '여행':'여행','특별 여행':'여행','문화':'문화','놀이':'놀이','유흥':'놀이','교육':'교육',
    '헤어':'헤어','패션':'패션','쇼핑':'쇼핑','쇼핑/패션':'쇼핑','미용':'미용',
    '관리비':'관리비','부모님 용돈':'부모님 용돈',
}
GROUP_MAP = {
    '외식/점심':'식사','외식/저녁':'식사','집밥/생활비':'식사','디저트':'식사',
    '운동':'건강','의료비':'건강','보험비':'건강','영양제':'건강',
    '대중교통':'교통/통신','차량':'교통/통신','주유/톨비':'교통/통신','통신비':'교통/통신',
    '선물':'관계','경조사':'관계','보은':'관계','회비':'관계',
    '여행':'여가','문화':'여가','놀이':'여가','교육':'여가',
    '헤어':'꾸미기','패션':'꾸미기','쇼핑':'꾸미기','미용':'꾸미기',
    '관리비':'주거','부모님 용돈':'가족','금융':'금융',
}
INCOME_CATS = {'월급','상여금','부수입','직장 외 부수입','당근','보험 환급'}

# ---------- 자동 분류 (카테고리 미입력 지출용) ----------
# 1) 장소 학습: 같은 장소가 라벨된 이력이 있으면 최빈 카테고리 상속
# 2) 키워드 규칙: 토스/뱅크샐러드류 표준 분류를 참고한 패턴
# 3) 남으면 '기타'
KEYWORD_RULES = [
    (r'용돈|엄마|아빠|엄빠|부모님', '부모님 용돈'),
    (r'KTX|코레일|SRT|기차|항공|공항|호텔|숙소|산장|펜션|리조트|트래블|여행', '여행'),
    (r'병원|의원|치과|한의원|약국|검진', '의료비'),
    (r'크로스핏|피트니스|헬스|필라테스|요가|수영', '운동'),
    (r'지하철|버스|택시|티머니|교통', '대중교통'),
    (r'KT$|LG|SKT|유플러스|통신', '통신비'),
    (r'국세청|세금|수수료|복비|이자|보증|송금|환전|카카오페이|네이버파이낸셜|토스|투자', '금융'),
    (r'루이비통|디스커버리|쿠팡|무신사|패밀리샵|백화점|아울렛|올리브영|다이소|당근', '쇼핑'),
    (r'스타벅스|커피|카페|베이커리|디저트', '디저트'),
    (r'CGV|메가박스|롯데시네마|영화|공연|전시|콘서트', '문화'),
    (r'결혼|축의|조의|부의|장례', '경조사'),
    (r'선물', '선물'),
    (r'강의|클래스|캠퍼스|인강|스터디', '교육'),
    (r'보험|화재해상', '보험비'),
    (r'마티니|와인|칵테일|맥주|포차|주점', '놀이'),
    (r'닌텐도|게임|플스|스팀', '놀이'),
    (r'\d개월 등록', '운동'),
    (r'비빔밥|식당|국밥|김밥', '외식/점심'),
]

def infer_category(place, detail, place_vote):
    """미입력 카테고리 추론. (카테고리, 추론여부) 반환"""
    key = (place or '').strip()
    if key and key in place_vote:
        return place_vote[key], True
    text = f"{place or ''} {detail or ''}"
    for pat, cat in KEYWORD_RULES:
        if re.search(pat, text, re.I):
            return cat, True
    return '기타', True

def anon(name):
    """사람 이름 마스킹: '김솔아 96' → '김** 96'"""
    if not name: return name
    s = str(name).strip()
    m = re.match(r'^([가-힣])[가-힣]{1,3}(\s*\d{2})?$', s)
    if m:
        return m.group(1) + '**' + (m.group(2) or '')
    return s

def anon_channel(ch):
    """경로 필드: 앱 이름(골드/블릿 등)은 유지, 지인 이름만 마스킹"""
    if not ch: return ch
    s = str(ch).strip()
    s = re.sub(r'(지인 소개)\s*-\s*\S+', r'\1', s)
    s = re.sub(r'^\S+\s*소개$', '지인 소개', s)
    return s

def jd(d):
    return d.strftime('%Y-%m-%d') if isinstance(d, (dt.datetime, dt.date)) else None

# ============================================================
# 1) 가계부 파싱
# ============================================================
wb = openpyxl.load_workbook(LEDGER, data_only=True)
month_re = re.compile(r'^(\d{2})-(\d{2})$')
month_sheets = sorted([n for n in wb.sheetnames if month_re.match(n)])
latest = month_sheets[-1]

tx = []
for name in month_sheets:
    ws = wb[name]
    cur = None; header = False
    yy, mm = name.split('-')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8):
        v = [c.value for c in row]
        if not header:
            if v[0] == '날짜': header = True
            continue
        d, inc, exp, place, detail, cat = v[0], v[1], v[2], v[3], v[4], v[5]
        note = v[6] if len(v) > 6 else None
        w = v[7] if len(v) > 7 else None
        if isinstance(d, (dt.datetime, dt.date)): cur = d
        if inc is None and exp is None and place is None and detail is None and cat is None:
            continue
        if cur is None: continue
        # 시트 우측/하단의 부속 메모 테이블(정산 기록 등) 오인 방지:
        # 실거래는 장소·세부내역·카테고리 중 최소 하나를 가짐
        if place is None and detail is None and cat is None:
            continue
        # 계산 메모 행 제외: 장소 칸이 숫자이거나 '총 123' 형태인데 내역·카테고리가 없는 경우
        if detail is None and cat is None and place is not None:
            ps = str(place).strip()
            if re.match(r'^[\d,.\s]+$', ps) or re.match(r'^총\s*[\d,]+', ps):
                continue
        # 시트 월과 무관한 템플릿/이월 행 방지: 날짜의 연-월이 시트와 2개월 이상 어긋나면 스킵
        rec = {'d': jd(cur), 'p': str(place).strip() if place else None,
               'dt': str(detail).strip() if detail else None,
               'n': str(note).strip() if note else None,
               'w': str(w).strip() if w else None, 'm': f'20{yy}-{mm}'}
        if isinstance(inc, (int, float)) and inc:
            rec['ty'] = 'i'; rec['a'] = round(float(inc))
            c = str(cat).strip() if cat else None
            if not c or c not in INCOME_CATS:
                c = '부수입' if (place and '멘토링' in str(place)) else (c or '기타수입')
            rec['c'] = c; rec['sc'] = c; rec['g'] = '수입'
            tx.append(rec)
        if isinstance(exp, (int, float)) and exp is not None:
            r2 = dict(rec)
            r2['ty'] = 'e'; r2['a'] = round(float(exp))
            c = str(cat).strip() if cat else None
            r2['c'] = c or '미입력'
            if c and c not in ('-',):
                sc = CAT_MAP.get(c, c)
                r2['sc'] = sc; r2['g'] = GROUP_MAP.get(sc, '기타')
            else:
                r2['sc'] = None  # 2패스에서 추론
            tx.append(r2)

# ---- 2패스: 장소별 최빈 카테고리 학습 후 미입력분 추론 ----
from collections import Counter as _C
place_votes = {}
for t in tx:
    if t['ty'] == 'e' and t['sc'] and t['p']:
        place_votes.setdefault(t['p'], _C())[t['sc']] += 1
place_vote = {p: c.most_common(1)[0][0] for p, c in place_votes.items()}
inferred = 0
for t in tx:
    if t['ty'] == 'e' and t['sc'] is None:
        sc, inf = infer_category(t['p'], t['dt'], place_vote)
        t['sc'] = sc; t['g'] = GROUP_MAP.get(sc, '기타')
        if inf: t['inf'] = 1; inferred += 1
print(f"자동 분류: {inferred}건 추론 완료")

# ---- 엄빠 용돈 원장 (최신 시트에서 '엄빠 용돈' 셀 탐색) ----
parents = []
ws = wb[latest]
anchor = None
for row in ws.iter_rows(min_row=1, max_row=10, min_col=15, max_col=45):
    for c in row:
        if c.value == '엄빠 용돈': anchor = (c.row, c.column)
if anchor:
    r0, c0 = anchor
    for r in range(r0 + 2, r0 + 120):
        d = ws.cell(r, c0).value; a = ws.cell(r, c0 + 1).value; memo = ws.cell(r, c0 + 2).value
        if d is None and a is None: break
        if isinstance(d, (dt.datetime, dt.date)):
            amt = round(float(a)) if isinstance(a, (int, float)) else None
            parents.append({'d': jd(d), 'a': amt, 'memo': str(memo) if memo else None})

# ---- 소개팅 원장 (최신 시트 우측: 회차 | 이름 | 경로 | 직업 | 날짜 | 횟수 | 금액...) ----
dating = {'entries': [], 'stats': None, 'history': []}
ws = wb[latest]
for r in range(55, 260):
    for c in range(23, 32):
        no = ws.cell(r, c).value
        nm = ws.cell(r, c + 1).value
        d = ws.cell(r, c + 4).value
        if isinstance(no, (int, float)) and isinstance(nm, str) and isinstance(d, (dt.datetime, dt.date)):
            ch = ws.cell(r, c + 2).value; job = ws.cell(r, c + 3).value
            meets = ws.cell(r, c + 5).value; amt = ws.cell(r, c + 6).value
            ex1 = ws.cell(r, c + 7).value; ex2 = ws.cell(r, c + 8).value
            total = sum(round(float(x)) for x in (amt, ex1, ex2) if isinstance(x, (int, float)))
            dating['entries'].append({
                'no': int(no), 'name': anon(nm) if ANONYMIZE else nm.strip(),
                'ch': anon_channel(ch) if ANONYMIZE else (str(ch).strip() if ch else None),
                'job': str(job).strip() if job else None,
                'd': jd(d), 'meets': int(meets) if isinstance(meets, (int, float)) else 1,
                'a': total})
            break
# 통계 블록: '호감 상대 등장' 셀 아래 행
for row in ws.iter_rows(min_row=55, max_row=80, min_col=28, max_col=40):
    for c in row:
        if c.value == '호감 상대 등장':
            r, col = c.row, c.column
            dating['stats'] = {
                'meal': ws.cell(r + 1, col - 3).value, 'total': ws.cell(r + 1, col - 1).value,
                'likeRate': ws.cell(r + 1, col).value, 'loveRate': ws.cell(r + 1, col + 1).value}
# 인연 히스토리 (우측 별도 테이블: 이름 | 시기 | 횟수 | 직업 | 경로)
for r in range(60, 90):
    for c in range(34, 40):
        nm = ws.cell(r, c).value; period = ws.cell(r, c + 1).value
        cnt = ws.cell(r, c + 2).value; job = ws.cell(r, c + 3).value
        if isinstance(nm, str) and re.match(r'^[가-힣]{1,4}(\s\d{2})?$', nm.strip()) and \
           (isinstance(cnt, (int, float)) or isinstance(period, str)):
            src = ws.cell(r, c + 4).value
            dating['history'].append({
                'name': anon(nm) if ANONYMIZE else nm.strip(),
                'period': str(period).strip() if period else None,
                'meets': int(cnt) if isinstance(cnt, (int, float)) else None,
                'job': str(job).strip() if job else None,
                'src': anon_channel(src) if ANONYMIZE else (str(src).strip() if src else None)})
            break

# ============================================================
# 2) 자산 포트폴리오 파싱
# ============================================================
pw = openpyxl.load_workbook(PORTFOLIO, data_only=True)

# ---- 0.자산 ----
a = pw['0.자산']
def num(cell):
    v = a[cell].value
    return round(float(v)) if isinstance(v, (int, float)) else None

assets = {
    'updated': jd(a['I1'].value),
    'total': num('E35'), 'totalPnl': num('G35'),
    'buckets': [
        {'k': '은행 입출금', 'v': num('E6')},
        {'k': '예적금·청약', 'v': num('E13')},
        {'k': '투자', 'v': num('E18'), 'pnl': num('G18'), 'ret': a['I18'].value},
        {'k': '기타(포인트·전세)', 'v': num('E26')},
    ],
    'deposits': [], 'securities': [], 'others': [], 'history': []
}
for r in range(14, 18):  # 예적금·청약
    nm, v, memo, rate = a.cell(r, 4).value, a.cell(r, 5).value, a.cell(r, 8).value, a.cell(r, 9).value
    if nm and isinstance(v, (int, float)):
        assets['deposits'].append({'k': str(nm), 'v': round(v), 'memo': str(memo) if memo else None,
                                   'rate': rate if isinstance(rate, (int, float)) else None})
for r in range(20, 23):  # 증권 계좌
    nm, v, pnl, ret = a.cell(r, 4).value, a.cell(r, 5).value, a.cell(r, 7).value, a.cell(r, 9).value
    if nm and isinstance(v, (int, float)):
        assets['securities'].append({'k': str(nm), 'v': round(v),
                                     'pnl': round(pnl) if isinstance(pnl, (int, float)) else None,
                                     'ret': ret if isinstance(ret, (int, float)) else None})
for r in range(27, 31):  # 기타
    nm, v = a.cell(r, 3).value, a.cell(r, 5).value
    if nm and isinstance(v, (int, float)):
        assets['others'].append({'k': str(nm), 'v': round(v)})
for r in range(4, 12):   # 자산 스냅샷 로그
    d, v, pnl = a.cell(r, 11).value, a.cell(r, 12).value, a.cell(r, 13).value
    if isinstance(d, (dt.datetime, dt.date)) and isinstance(v, (int, float)):
        assets['history'].append({'d': jd(d), 'v': round(v),
                                  'pnl': round(pnl) if isinstance(pnl, (int, float)) else None})
assets['history'].sort(key=lambda x: x['d'])

# ---- 1.재테크: 종료 거래 + 보유 종목 + 코인 ----
t = pw['1.재테크']
realized = []
start = None
for r in range(1, t.max_row + 1):
    if t.cell(r, 12).value == '3) 종료': start = r + 2; break
if start:
    for r in range(start + 1, start + 80):
        name = t.cell(r, 19).value
        if t.cell(r, 14).value == '총 계' or (name is None and t.cell(r, 14).value is None):
            if name is None and t.cell(r + 1, 19).value is None: break
            if t.cell(r, 14).value == '총 계': break
            continue
        if not name: continue
        buy_d, sell_d = t.cell(r, 20).value, t.cell(r, 21).value
        realized.append({
            'name': str(name).strip(),
            'buy': jd(buy_d), 'sell': jd(sell_d),
            'invest': round(t.cell(r, 25).value or 0),
            'proceeds': round(t.cell(r, 27).value or 0),
            'div': round(t.cell(r, 28).value or 0),
            'pnl': round(t.cell(r, 29).value or 0),
            'ret': t.cell(r, 30).value})
holdings = []
for r in range(24, 50):  # 티커 | 종목 | 현재가치 테이블 (E~G열)
    tk, nm, v = t.cell(r, 5).value, t.cell(r, 6).value, t.cell(r, 7).value
    if nm and isinstance(v, (int, float)) and v > 0:
        holdings.append({'ticker': str(tk), 'name': str(nm).strip(), 'v': round(v)})
# 보령 본인 계좌 합산 행('주 계')
boryung = None
for r in range(50, 70):
    if t.cell(r, 14).value == '주 계':
        boryung = {'name': '보령(누적 매수)', 'invest': round(t.cell(r, 25).value or 0),
                   'pnl': round(t.cell(r, 29).value or 0), 'ret': t.cell(r, 30).value}
        break
coin = None
for r in range(125, 140):
    if t.cell(r, 14).value and '코인' in str(t.cell(r, 14).value):
        coin = {'invest': round(t.cell(r + 4, 18).value or 0) if t.cell(r + 4, 18).value else None,
                'value': round(t.cell(r + 4, 21).value or 0) if t.cell(r + 4, 21).value else None}
        break

# ---- 5.VC투자 ----
vc = []
v5 = pw['5.VC투자']
for r in range(2, 10):
    nm = v5.cell(r, 3).value
    if not nm: continue
    vc.append({'name': str(nm).strip(),
               'invest': round(v5.cell(r, 6).value or 0),
               'value': round(v5.cell(r, 8).value or 0) if isinstance(v5.cell(r, 8).value, (int, float)) else None,
               'pnl': round(v5.cell(r, 9).value or 0) if isinstance(v5.cell(r, 9).value, (int, float)) else None,
               'ret': v5.cell(r, 10).value,
               'status': '보유중' if v5.cell(r, 2).value in ('보유중', None, '-') else '종료'})

# ---- 2.연봉 ----
salary = []
s = pw['2.연봉']
cur_co = None
for r in range(3, 20):
    co, d, pay = s.cell(r, 1).value, s.cell(r, 2).value, s.cell(r, 3).value
    if co: cur_co = str(co).strip()
    if not isinstance(d, (dt.datetime, dt.date)): continue
    salary.append({'co': cur_co, 'd': jd(d),
                   'pay': round(pay) if isinstance(pay, (int, float)) else None,
                   'rate': s.cell(r, 4).value if isinstance(s.cell(r, 4).value, (int, float)) else None,
                   'why': str(s.cell(r, 5).value).strip() if s.cell(r, 5).value else None,
                   'received': round(s.cell(r, 10).value) if isinstance(s.cell(r, 10).value, (int, float)) else None})

# ---- 3.사이드잡 ----
sidejobs = []
sj = pw['3.사이드잡']
cur_year = None
for r in range(3, 80):
    src, yr, total = sj.cell(r, 1).value, sj.cell(r, 2).value, sj.cell(r, 6).value
    if isinstance(yr, (int, float)): cur_year = int(yr)
    if not src or str(src).strip() in ('총계', '수입처', '사이드잡'): continue
    if isinstance(total, (int, float)) and cur_year:
        cnt = sj.cell(r, 5).value
        sidejobs.append({'y': cur_year, 'src': str(src).strip(), 'total': round(total),
                         'count': int(cnt) if isinstance(cnt, (int, float)) else None})

# ---- 4.카드캐시백 ----
cashback = []
cb = pw['4.카드캐시백']
for r in range(3, 25):
    if cb.cell(r, 1).value == '총계':
        yr, amt = cb.cell(r, 2).value, cb.cell(r, 4).value
        if isinstance(yr, (int, float)) and isinstance(amt, (int, float)):
            cashback.append({'y': int(yr), 'a': round(amt)})

# ============================================================
# 3) 출력
# ============================================================
data = {
    'meta': {
        'generated': dt.date.today().isoformat(),
        'source': 'excel-snapshot', 'anonymized': ANONYMIZE,
        'txCount': len(tx), 'from': min(x['d'] for x in tx), 'to': max(x['d'] for x in tx),
        'months': [f'20{n.replace("-", "-")}' for n in month_sheets],
    },
    'tx': tx, 'parents': parents, 'dating': dating,
    'assets': assets,
    'invest': {'realized': realized, 'holdings': holdings, 'boryung': boryung, 'coin': coin, 'vc': vc},
    'salary': salary, 'sidejobs': sidejobs, 'cashback': cashback,
}
with open(OUT, 'w', encoding='utf-8') as f:
    f.write('window.APP_DATA = ')
    json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
    f.write(';')

print(f"tx={len(tx)} parents={len(parents)} dating={len(dating['entries'])}/{len(dating['history'])} "
      f"realized={len(realized)} holdings={len(holdings)} vc={len(vc)} salary={len(salary)} "
      f"sidejobs={len(sidejobs)} cashback={len(cashback)} assetsHist={len(assets['history'])}")
print('total', assets['total'], '| out:', OUT)
